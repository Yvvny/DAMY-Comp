from __future__ import annotations

import csv
import html
import hashlib
import os
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process

from folder_manager.sms.cloudflare_r2 import missing_r2_settings, upload_file_to_r2
from folder_manager.sms.twilio_client import normalize_us_phone, send_sms

CHILD_INFO_FILENAME = "Child Info.xlsx"
PARENT_CONTACTS_DISPLAY_NAME = "Parent Contacts"
CHILD_ASSETS_DIRNAME = "Parent Delivery Assets"
VISIBLE_SHEET_TITLE = "Parent Contacts"
INTERNAL_SHEET_TITLE = "_internal"

DEFAULT_PARENT_MMS_BODY = (
    "Hello, your child's school photo proofs are ready. "
    "Scan the QR code in the attached image to view and order. Reply STOP to opt out."
)
DEFAULT_MMS_BODY = DEFAULT_PARENT_MMS_BODY
DEFAULT_PARENT_EMAIL_SUBJECT = "Your child's school photo proofs are ready"

STATUS_READY = "Ready"
STATUS_MISSING = "Missing"
STATUS_NEEDS_REVIEW = "Needs review"
STATUS_SENT = "Sent"
STATUS_FAILED = "Failed"

VISIBLE_HEADERS = [
    "Child Name",
    "Class",
    "Parent Phone",
    "Parent Email",
    "MMS",
    "Email",
    "Note",
]
INTERNAL_HEADERS = [
    "Page Index",
    "Password",
    "Preview URL",
    "Local Image Path",
    "Source PDF Path",
    "Order URL",
    "Object Key",
    "MMS SID",
    "MMS Sent At",
    "MMS Error",
    "Email Message ID",
    "Email Sent At",
    "Email Error",
]

EMAIL_RE = re.compile(r"([A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,})", re.IGNORECASE)
PHONE_RE = re.compile(r"(?:\+?1[\s\-.]?)?(?:\(?\d{3}\)?[\s\-.]?)\d{3}[\s\-.]?\d{4}")
GENERIC_CONTACT_FIELDS = {
    "name",
    "child name",
    "student name",
    "first name",
    "last name",
    "firstname",
    "lastname",
    "phone",
    "phone number",
    "parent phone",
    "mobile",
    "cell",
    "email",
    "email address",
    "parent email",
    "class",
    "grade",
    "contact",
    "parent",
    "birthdate",
    "date of birth",
    "dob",
}

_PARENT_DELIVERY_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS parent_delivery_contacts (
    id                BIGSERIAL PRIMARY KEY,
    workflow_run_id   BIGINT NOT NULL REFERENCES workflow_runs(id) ON DELETE CASCADE,
    job_disk_name     TEXT NOT NULL DEFAULT '',
    child_name        TEXT NOT NULL DEFAULT '',
    class_name        TEXT NOT NULL DEFAULT '',
    page_index        INTEGER NOT NULL DEFAULT 0,
    password          TEXT NOT NULL DEFAULT '',
    parent_phone      TEXT NOT NULL DEFAULT '',
    parent_email      TEXT NOT NULL DEFAULT '',
    note              TEXT NOT NULL DEFAULT '',
    preview_url       TEXT NOT NULL DEFAULT '',
    local_image_path  TEXT NOT NULL DEFAULT '',
    source_pdf_path   TEXT NOT NULL DEFAULT '',
    order_url         TEXT NOT NULL DEFAULT '',
    object_key        TEXT NOT NULL DEFAULT '',
    mms_sid           TEXT NOT NULL DEFAULT '',
    mms_sent_at       TEXT NOT NULL DEFAULT '',
    mms_error         TEXT NOT NULL DEFAULT '',
    email_message_id  TEXT NOT NULL DEFAULT '',
    email_sent_at     TEXT NOT NULL DEFAULT '',
    email_error       TEXT NOT NULL DEFAULT '',
    created_at        TIMESTAMPTZ NOT NULL DEFAULT now(),
    updated_at        TIMESTAMPTZ NOT NULL DEFAULT now()
);

ALTER TABLE parent_delivery_contacts
ADD COLUMN IF NOT EXISTS page_index INTEGER NOT NULL DEFAULT 0;

ALTER TABLE parent_delivery_contacts
ADD COLUMN IF NOT EXISTS password TEXT NOT NULL DEFAULT '';

ALTER TABLE parent_delivery_contacts
ADD COLUMN IF NOT EXISTS order_url TEXT NOT NULL DEFAULT '';

CREATE UNIQUE INDEX IF NOT EXISTS idx_parent_delivery_item_pdf_page
ON parent_delivery_contacts(workflow_run_id, source_pdf_path, page_index);

CREATE INDEX IF NOT EXISTS idx_parent_delivery_item_status
ON parent_delivery_contacts(workflow_run_id, child_name, class_name);
"""


@dataclass
class ChildRecord:
    child_name: str
    class_name: str
    db_id: int = 0
    workflow_item_id: int = 0
    page_index: int = 0
    password: str = ""
    parent_phone: str = ""
    parent_email: str = ""
    mms_status: str = STATUS_MISSING
    email_status: str = STATUS_MISSING
    note: str = ""
    preview_url: str = ""
    local_image_path: str = ""
    source_pdf_path: str = ""
    order_url: str = ""
    object_key: str = ""
    mms_sid: str = ""
    mms_sent_at: str = ""
    mms_error: str = ""
    email_message_id: str = ""
    email_sent_at: str = ""
    email_error: str = ""


@dataclass
class ContactEntry:
    name: str
    phone: str = ""
    email: str = ""


def _sanitize_excel_text(value: object) -> str:
    text = str(value or "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    return ILLEGAL_CHARACTERS_RE.sub("", text).strip()


def _now_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %I:%M %p")


def _safe_slug(text: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "-", str(text or "").lower())
    return re.sub(r"-+", "-", value).strip("-") or "job"


def _normalize_class_for_key(text: str) -> str:
    value = _sanitize_excel_text(text)
    value = re.sub(r"\bpdfs\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\bpassword\b", "", value, flags=re.IGNORECASE)
    value = re.sub(r"\s+", " ", value).strip(" -_")
    return value.casefold()


def _normalize_name_key(text: str) -> str:
    value = _sanitize_excel_text(text).casefold()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return " ".join(value.split())


def _record_key(child_name: str, class_name: str) -> str:
    return f"{_normalize_name_key(child_name)}|{_normalize_class_for_key(class_name)}"


def _note_indicates_review(note: str) -> bool:
    lowered = _sanitize_excel_text(note).casefold()
    return lowered.startswith("needs review") or lowered.startswith("review:")


def _strip_note_prefix(note: str, prefixes: Sequence[str]) -> str:
    lines = [line for line in _sanitize_excel_text(note).splitlines() if line.strip()]
    kept = []
    lowered_prefixes = tuple(prefix.casefold() for prefix in prefixes)
    for line in lines:
        lowered = line.casefold()
        if any(lowered.startswith(prefix) for prefix in lowered_prefixes):
            continue
        kept.append(line)
    return "\n".join(kept)


def _append_note(note: str, message: str, *, replace_prefixes: Sequence[str] = ()) -> str:
    base = _strip_note_prefix(note, replace_prefixes) if replace_prefixes else _sanitize_excel_text(note)
    extra = _sanitize_excel_text(message)
    if not extra:
        return base
    if not base:
        return extra
    return f"{base}\n{extra}"


def _clean_contact_name(value: str) -> str:
    text = _sanitize_excel_text(value)
    text = PHONE_RE.sub(" ", text)
    text = EMAIL_RE.sub(" ", text)
    text = text.strip(" -|,:;")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _extract_email(text: str) -> str:
    match = EMAIL_RE.search(str(text or ""))
    return _sanitize_excel_text(match.group(1)).lower() if match else ""


def _extract_phone(text: str) -> str:
    match = PHONE_RE.search(str(text or ""))
    if not match:
        return ""
    value = _sanitize_excel_text(match.group(0))
    try:
        return normalize_us_phone(value)
    except Exception:
        digits = re.sub(r"\D+", "", value)
        return digits


def validate_parent_phone(value: str) -> str:
    text = _sanitize_excel_text(value)
    if not text:
        return ""
    try:
        return normalize_us_phone(text)
    except Exception as exc:
        raise ValueError("Use a valid US phone number, for example 9297349818.") from exc


def validate_parent_email(value: str) -> str:
    text = _sanitize_excel_text(value).lower()
    if not text:
        return ""
    if not EMAIL_RE.fullmatch(text):
        raise ValueError("Use a valid email address, for example parent@example.com.")
    return text


def child_info_path(job_folder: str | Path) -> Path:
    return Path(job_folder).resolve() / CHILD_INFO_FILENAME


def child_assets_path(job_folder: str | Path) -> Path:
    return _pdfs_root(job_folder) / CHILD_ASSETS_DIRNAME


def _pdfs_root(job_folder: str | Path) -> Path:
    return Path(job_folder).resolve() / "PDFs"


def _collect_pdf_files(root: Path) -> list[Path]:
    if not root.is_dir():
        return []
    files = [
        path
        for path in root.rglob("*.pdf")
        if path.is_file() and CHILD_ASSETS_DIRNAME.casefold() not in {part.casefold() for part in path.parts}
    ]
    files.sort(key=lambda item: str(item).lower())
    return files


def _class_name_from_pdf_path(pdf_path: Path) -> str:
    parent_name = pdf_path.parent.name if pdf_path.parent else ""
    if parent_name.lower().endswith(" pdfs"):
        parent_name = parent_name[:-5]
    parent_name = re.sub(r"\s+", " ", parent_name).strip()
    return _sanitize_excel_text(parent_name)


def _extract_page_title_line(text: str) -> str:
    lines = [_sanitize_excel_text(line) for line in str(text or "").splitlines()]
    filtered = [line for line in lines if line]
    ignore_prefixes = (
        "order now",
        "password:",
        "it'",
        "aun",
        "view & order",
    )
    for line in filtered:
        lowered = line.casefold()
        if lowered.startswith(ignore_prefixes):
            continue
        return line
    return ""


def _extract_page_password(text: str) -> str:
    for line in str(text or "").splitlines():
        cleaned = _sanitize_excel_text(line)
        match = re.search(r"\b(?:password|passcode)\b\s*[:#-]?\s*([A-Za-z0-9]{3,32})", cleaned, re.IGNORECASE)
        if match:
            return _sanitize_excel_text(match.group(1))
    match = re.search(r"\b(?:password|passcode)\b\s*[:#-]?\s*([A-Za-z0-9]{3,32})", str(text or ""), re.IGNORECASE)
    return _sanitize_excel_text(match.group(1)) if match else ""


def _parse_page_identity(text: str, pdf_path: Path, page_index: int) -> tuple[str, str]:
    title_line = _extract_page_title_line(text)
    class_name = _class_name_from_pdf_path(pdf_path)
    child_name = ""

    if title_line:
        match = re.match(r"^(?P<child>.+?)\s*-\s*(?P<class>Class\b.+)$", title_line, re.IGNORECASE)
        if match:
            child_name = _sanitize_excel_text(match.group("child"))
            class_name = _sanitize_excel_text(match.group("class")) or class_name
        else:
            child_name = title_line

    if not child_name:
        child_name = f"Page {page_index + 1}"
    if not class_name:
        class_name = "Unknown Class"
    return child_name, class_name


def _public_url(base_url: str, object_key: str) -> str:
    return f"{str(base_url or '').rstrip('/')}/{str(object_key).lstrip('/')}"


def _stable_object_key(job_folder: str | Path, pdf_path: Path, page_index: int, child_name: str, class_name: str) -> str:
    job_name = Path(job_folder).resolve().name
    digest = hashlib.sha1(
        f"{pdf_path.as_posix()}|{page_index}|{child_name}|{class_name}".encode("utf-8", errors="ignore")
    ).hexdigest()[:24]
    return f"parent-delivery/{_safe_slug(job_name)}/{digest}.jpg"


def _stable_asset_stem(pdf_path: Path, page_index: int, child_name: str, class_name: str) -> str:
    digest = hashlib.sha1(
        f"{pdf_path.as_posix()}|{page_index}|{child_name}|{class_name}".encode("utf-8", errors="ignore")
    ).hexdigest()[:10]
    class_part = _safe_slug(class_name or "class")
    child_part = _safe_slug(child_name or f"page-{page_index + 1}")
    return f"{class_part} - {child_part} - p{page_index + 1} - {digest}"


def _extract_order_url_from_page(page) -> str:
    try:
        links = page.get_links()
    except Exception:
        links = []
    for link in links or []:
        uri = str((link or {}).get("uri") or "").strip()
        if uri.startswith(("http://", "https://")):
            return uri
    return ""


def _render_page_preview(page, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    pix = page.get_pixmap(matrix=page.parent[page.number].get_pixmap().matrix if False else None)
    if pix.alpha:
        pix = page.get_pixmap(alpha=False)
    pix.save(str(target_path))


def _db_mode(db: Any = None, workflow_item_id: Optional[int] = None) -> bool:
    try:
        return db is not None and int(workflow_item_id or 0) > 0
    except Exception:
        return False


def _ensure_parent_delivery_schema(db: Any) -> None:
    if getattr(db, "_parent_delivery_schema_ready", False):
        return
    with db.connect() as conn, conn.cursor() as cur:
        cur.execute(_PARENT_DELIVERY_SCHEMA_SQL)
    try:
        setattr(db, "_parent_delivery_schema_ready", True)
    except Exception:
        pass


def _row_to_child_record(row: Sequence[object]) -> ChildRecord:
    record = ChildRecord(
        db_id=int(row[0] or 0),
        workflow_item_id=int(row[1] or 0),
        child_name=_sanitize_excel_text(row[2]),
        class_name=_sanitize_excel_text(row[3]),
        page_index=int(row[4] or 0),
        password=_sanitize_excel_text(row[5]),
        parent_phone=_sanitize_excel_text(row[6]),
        parent_email=_sanitize_excel_text(row[7]).lower(),
        note=_sanitize_excel_text(row[8]),
        preview_url=_sanitize_excel_text(row[9]),
        local_image_path=_sanitize_excel_text(row[10]),
        source_pdf_path=_sanitize_excel_text(row[11]),
        order_url=_sanitize_excel_text(row[12]),
        object_key=_sanitize_excel_text(row[13]),
        mms_sid=_sanitize_excel_text(row[14]),
        mms_sent_at=_sanitize_excel_text(row[15]),
        mms_error=_sanitize_excel_text(row[16]),
        email_message_id=_sanitize_excel_text(row[17]),
        email_sent_at=_sanitize_excel_text(row[18]),
        email_error=_sanitize_excel_text(row[19]),
    )
    record.mms_status = _compute_mms_status(record)
    record.email_status = _compute_email_status(record)
    return record


def _read_child_records_from_db(db: Any, workflow_item_id: int) -> list[ChildRecord]:
    _ensure_parent_delivery_schema(db)
    with db.connect() as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT
                id, workflow_run_id, child_name, class_name, page_index,
                password, parent_phone, parent_email, note, preview_url, local_image_path,
                source_pdf_path, order_url, object_key, mms_sid, mms_sent_at,
                mms_error, email_message_id, email_sent_at, email_error
            FROM parent_delivery_contacts
            WHERE workflow_run_id=%s
            ORDER BY lower(class_name), lower(child_name), page_index, id
            """,
            (int(workflow_item_id),),
        )
        rows = cur.fetchall()
    return [_row_to_child_record(row) for row in rows]


def _write_child_records_to_db(
    db: Any,
    workflow_item_id: int,
    records: Sequence[ChildRecord],
    *,
    disk_name: str = "",
) -> None:
    _ensure_parent_delivery_schema(db)
    with db.connect() as conn, conn.cursor() as cur:
        for record in records:
            record.workflow_item_id = int(workflow_item_id)
            record.parent_phone = validate_parent_phone(record.parent_phone)
            record.parent_email = validate_parent_email(record.parent_email)
            record.mms_status = _compute_mms_status(record)
            record.email_status = _compute_email_status(record)
            values = (
                int(workflow_item_id),
                _sanitize_excel_text(disk_name),
                _sanitize_excel_text(record.child_name),
                _sanitize_excel_text(record.class_name),
                int(record.page_index or 0),
                _sanitize_excel_text(record.password),
                _sanitize_excel_text(record.parent_phone),
                _sanitize_excel_text(record.parent_email).lower(),
                _sanitize_excel_text(record.note),
                _sanitize_excel_text(record.preview_url),
                _sanitize_excel_text(record.local_image_path),
                _sanitize_excel_text(record.source_pdf_path),
                _sanitize_excel_text(record.order_url),
                _sanitize_excel_text(record.object_key),
                _sanitize_excel_text(record.mms_sid),
                _sanitize_excel_text(record.mms_sent_at),
                _sanitize_excel_text(record.mms_error),
                _sanitize_excel_text(record.email_message_id),
                _sanitize_excel_text(record.email_sent_at),
                _sanitize_excel_text(record.email_error),
            )
            if int(record.db_id or 0) > 0:
                cur.execute(
                    """
                    UPDATE parent_delivery_contacts
                    SET
                        job_disk_name=%s,
                        child_name=%s,
                        class_name=%s,
                        page_index=%s,
                        password=%s,
                        parent_phone=%s,
                        parent_email=%s,
                        note=%s,
                        preview_url=%s,
                        local_image_path=%s,
                        source_pdf_path=%s,
                        order_url=%s,
                        object_key=%s,
                        mms_sid=%s,
                        mms_sent_at=%s,
                        mms_error=%s,
                        email_message_id=%s,
                        email_sent_at=%s,
                        email_error=%s,
                        updated_at=now()
                    WHERE id=%s AND workflow_run_id=%s
                    """,
                    values[1:] + (int(record.db_id), int(workflow_item_id)),
                )
                continue
            cur.execute(
                """
                INSERT INTO parent_delivery_contacts(
                    workflow_run_id, job_disk_name, child_name, class_name, page_index,
                    password, parent_phone, parent_email, note, preview_url, local_image_path,
                    source_pdf_path, order_url, object_key, mms_sid, mms_sent_at,
                    mms_error, email_message_id, email_sent_at, email_error
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (workflow_run_id, source_pdf_path, page_index) DO UPDATE SET
                    job_disk_name=EXCLUDED.job_disk_name,
                    child_name=EXCLUDED.child_name,
                    class_name=EXCLUDED.class_name,
                    password=EXCLUDED.password,
                    parent_phone=EXCLUDED.parent_phone,
                    parent_email=EXCLUDED.parent_email,
                    note=EXCLUDED.note,
                    preview_url=EXCLUDED.preview_url,
                    local_image_path=EXCLUDED.local_image_path,
                    order_url=EXCLUDED.order_url,
                    object_key=EXCLUDED.object_key,
                    mms_sid=EXCLUDED.mms_sid,
                    mms_sent_at=EXCLUDED.mms_sent_at,
                    mms_error=EXCLUDED.mms_error,
                    email_message_id=EXCLUDED.email_message_id,
                    email_sent_at=EXCLUDED.email_sent_at,
                    email_error=EXCLUDED.email_error,
                    updated_at=now()
                RETURNING id
                """,
                values,
            )
            row = cur.fetchone()
            if row:
                record.db_id = int(row[0] or 0)


def _delete_stale_child_records_from_db(
    db: Any,
    workflow_item_id: int,
    records: Sequence[ChildRecord],
) -> None:
    keep_ids = [int(record.db_id or 0) for record in records if int(record.db_id or 0) > 0]
    _ensure_parent_delivery_schema(db)
    with db.connect() as conn, conn.cursor() as cur:
        if keep_ids:
            cur.execute(
                """
                DELETE FROM parent_delivery_contacts
                WHERE workflow_run_id=%s
                  AND NOT (id = ANY(%s))
                """,
                (int(workflow_item_id), keep_ids),
            )
        else:
            cur.execute(
                "DELETE FROM parent_delivery_contacts WHERE workflow_run_id=%s",
                (int(workflow_item_id),),
            )


def _record_page_key(record: ChildRecord) -> tuple[str, int]:
    return (_sanitize_excel_text(record.source_pdf_path), int(record.page_index or 0))


def _build_existing_maps(records: Sequence[ChildRecord]) -> tuple[Dict[tuple[str, int], ChildRecord], Dict[str, ChildRecord]]:
    by_page: Dict[tuple[str, int], ChildRecord] = {}
    by_name: Dict[str, ChildRecord] = {}
    for record in records:
        page_key = _record_page_key(record)
        if page_key[0]:
            by_page[page_key] = record
        name_key = _record_key(record.child_name, record.class_name)
        if name_key not in by_name:
            by_name[name_key] = record
    return by_page, by_name


def _merge_existing_fields(new_record: ChildRecord, existing: Optional[ChildRecord]) -> ChildRecord:
    if existing is None:
        return new_record
    new_record.db_id = existing.db_id or 0
    new_record.workflow_item_id = existing.workflow_item_id or new_record.workflow_item_id
    new_record.password = new_record.password or existing.password or ""
    new_record.parent_phone = existing.parent_phone or ""
    new_record.parent_email = existing.parent_email or ""
    new_record.note = existing.note or ""
    new_record.mms_sid = existing.mms_sid or ""
    new_record.mms_sent_at = existing.mms_sent_at or ""
    new_record.mms_error = existing.mms_error or ""
    new_record.email_message_id = existing.email_message_id or ""
    new_record.email_sent_at = existing.email_sent_at or ""
    new_record.email_error = existing.email_error or ""
    return new_record


def _load_existing_records(
    job_folder: str | Path,
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
) -> list[ChildRecord]:
    records = read_child_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    if records or not _db_mode(db, workflow_item_id):
        return records
    # One-time compatibility path: preserve contacts from an older Child Info.xlsx
    # before moving the job to DB-backed Parent Delivery.
    return read_child_records(job_folder)


def prepare_child_info_assets(
    job_folder: str | Path,
    r2_settings: Dict[str, str],
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Dict[str, object]:
    missing = missing_r2_settings(r2_settings)
    if missing:
        raise RuntimeError("Missing Cloudflare R2 settings: " + ", ".join(missing))

    job_folder = Path(job_folder).resolve()
    pdf_root = _pdfs_root(job_folder)
    if not pdf_root.is_dir():
        raise RuntimeError(f"PDFs folder not found: {pdf_root}")

    try:
        import fitz  # type: ignore
    except ImportError as exc:
        raise RuntimeError("PyMuPDF is required to build parent delivery previews.") from exc

    pdf_files = _collect_pdf_files(pdf_root)
    if not pdf_files:
        raise RuntimeError(f"No PDF files found under: {pdf_root}")

    assets_dir = child_assets_path(job_folder)
    existing_records = _load_existing_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    existing_by_page, existing_by_name = _build_existing_maps(existing_records)

    records: list[ChildRecord] = []
    uploaded_count = 0
    if assets_dir.exists():
        shutil.rmtree(assets_dir, ignore_errors=True)
    assets_dir.mkdir(parents=True, exist_ok=True)

    for pdf_path in pdf_files:
        with fitz.open(str(pdf_path)) as doc:
            for page_index in range(doc.page_count):
                page = doc.load_page(page_index)
                page_text = page.get_text("text")
                child_name, class_name = _parse_page_identity(page_text, pdf_path, page_index)
                object_key = _stable_object_key(job_folder, pdf_path, page_index, child_name, class_name)
                asset_stem = _stable_asset_stem(pdf_path, page_index, child_name, class_name)
                image_path = assets_dir / f"{asset_stem}.jpg"
                child_pdf_path = assets_dir / f"{asset_stem}.pdf"

                page.get_pixmap(alpha=False).save(str(image_path))
                child_doc = fitz.open()
                try:
                    child_doc.insert_pdf(doc, from_page=page_index, to_page=page_index)
                    child_doc.save(str(child_pdf_path))
                finally:
                    child_doc.close()

                upload_file_to_r2(
                    account_id=str(r2_settings.get("account_id") or ""),
                    access_key_id=str(r2_settings.get("access_key_id") or ""),
                    secret_access_key=str(r2_settings.get("secret_access_key") or ""),
                    bucket_name=str(r2_settings.get("bucket_name") or ""),
                    object_key=object_key,
                    file_path=str(image_path),
                    content_type="image/jpeg",
                )
                uploaded_count += 1
                record = ChildRecord(
                    workflow_item_id=int(workflow_item_id or 0),
                    child_name=child_name,
                    class_name=class_name,
                    page_index=page_index,
                    password=_extract_page_password(page_text),
                    preview_url=_public_url(str(r2_settings.get("public_base_url") or ""), object_key),
                    local_image_path=str(image_path),
                    source_pdf_path=str(child_pdf_path),
                    order_url=_extract_order_url_from_page(page),
                    object_key=object_key,
                )
                existing = existing_by_page.get((str(child_pdf_path), page_index)) or existing_by_name.get(
                    _record_key(child_name, class_name)
                )
                record = _merge_existing_fields(record, existing)
                if not record.order_url:
                    record.note = _append_note(
                        record.note,
                        "Needs review: PDF page does not contain an order link.",
                        replace_prefixes=("Needs review: PDF page does not contain an order link.",),
                    )
                records.append(record)

    write_child_records(job_folder, records, db=db, workflow_item_id=workflow_item_id, disk_name=disk_name)
    if _db_mode(db, workflow_item_id):
        _delete_stale_child_records_from_db(db, int(workflow_item_id or 0), records)
    return {
        "excel_path": "" if _db_mode(db, workflow_item_id) else str(child_info_path(job_folder)),
        "record_count": len(records),
        "uploaded_count": uploaded_count,
        "assets_dir": str(assets_dir),
    }


def _sheet_from_workbook(workbook) -> object:
    if VISIBLE_SHEET_TITLE in workbook.sheetnames:
        return workbook[VISIBLE_SHEET_TITLE]
    return workbook[workbook.sheetnames[0]]


def _hidden_sheet_from_workbook(workbook) -> Optional[object]:
    if INTERNAL_SHEET_TITLE in workbook.sheetnames:
        return workbook[INTERNAL_SHEET_TITLE]
    return None


def _coerce_record_from_rows(visible_map: Dict[str, object], internal_map: Dict[str, object]) -> ChildRecord:
    record = ChildRecord(
        child_name=_sanitize_excel_text(visible_map.get("Child Name")),
        class_name=_sanitize_excel_text(visible_map.get("Class")),
        page_index=int(_sanitize_excel_text(internal_map.get("Page Index")) or 0),
        password=_sanitize_excel_text(internal_map.get("Password")),
        parent_phone=_sanitize_excel_text(visible_map.get("Parent Phone")),
        parent_email=_sanitize_excel_text(visible_map.get("Parent Email")),
        note=_sanitize_excel_text(visible_map.get("Note")),
        preview_url=_sanitize_excel_text(internal_map.get("Preview URL")),
        local_image_path=_sanitize_excel_text(internal_map.get("Local Image Path")),
        source_pdf_path=_sanitize_excel_text(internal_map.get("Source PDF Path")),
        order_url=_sanitize_excel_text(internal_map.get("Order URL")),
        object_key=_sanitize_excel_text(internal_map.get("Object Key")),
        mms_sid=_sanitize_excel_text(internal_map.get("MMS SID") or internal_map.get("Twilio SID")),
        mms_sent_at=_sanitize_excel_text(internal_map.get("MMS Sent At") or internal_map.get("Sent At")),
        mms_error=_sanitize_excel_text(internal_map.get("MMS Error") or internal_map.get("Error")),
        email_message_id=_sanitize_excel_text(internal_map.get("Email Message ID")),
        email_sent_at=_sanitize_excel_text(internal_map.get("Email Sent At")),
        email_error=_sanitize_excel_text(internal_map.get("Email Error")),
    )
    legacy_status = _sanitize_excel_text(visible_map.get("Status"))
    if legacy_status and not record.mms_sent_at and legacy_status == STATUS_SENT:
        record.mms_sent_at = _now_stamp()
    record.parent_phone = _sanitize_excel_text(record.parent_phone)
    record.parent_email = _sanitize_excel_text(record.parent_email).lower()
    record.mms_status = _compute_mms_status(record)
    record.email_status = _compute_email_status(record)
    return record


def read_child_records(
    job_folder: str | Path,
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
) -> list[ChildRecord]:
    if _db_mode(db, workflow_item_id):
        return _read_child_records_from_db(db, int(workflow_item_id or 0))

    path = child_info_path(job_folder)
    if not path.is_file():
        return []

    workbook = load_workbook(path)
    visible_sheet = _sheet_from_workbook(workbook)
    hidden_sheet = _hidden_sheet_from_workbook(workbook)

    visible_headers = [_sanitize_excel_text(value) for value in next(visible_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    visible_index = {header: idx for idx, header in enumerate(visible_headers) if header}

    hidden_rows: list[Dict[str, object]] = []
    if hidden_sheet is not None:
        hidden_header_values = [_sanitize_excel_text(value) for value in next(hidden_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        hidden_index = {header: idx for idx, header in enumerate(hidden_header_values) if header}
        for row in hidden_sheet.iter_rows(min_row=2, values_only=True):
            hidden_rows.append({header: row[idx] if idx < len(row) else "" for header, idx in hidden_index.items()})

    records: list[ChildRecord] = []
    for row_number, row in enumerate(visible_sheet.iter_rows(min_row=2, values_only=True), start=0):
        visible_map = {header: row[idx] if idx < len(row) else "" for header, idx in visible_index.items()}
        internal_map = hidden_rows[row_number] if row_number < len(hidden_rows) else {}
        record = _coerce_record_from_rows(visible_map, internal_map)
        if not record.child_name and not record.class_name and not record.parent_phone and not record.parent_email:
            continue
        records.append(record)
    return records


def _compute_channel_status(*, contact_value: str, sent_at: str, error: str, note: str, asset_ready: bool) -> str:
    if not _sanitize_excel_text(contact_value):
        return STATUS_MISSING
    if _sanitize_excel_text(sent_at):
        return STATUS_SENT
    if _sanitize_excel_text(error):
        return STATUS_FAILED
    if _note_indicates_review(note):
        return STATUS_NEEDS_REVIEW
    if not asset_ready:
        return STATUS_FAILED
    return STATUS_READY


def _compute_mms_status(record: ChildRecord) -> str:
    return _compute_channel_status(
        contact_value=record.parent_phone,
        sent_at=record.mms_sent_at,
        error=record.mms_error,
        note=record.note,
        asset_ready=bool(record.preview_url),
    )


def _compute_email_status(record: ChildRecord) -> str:
    source_pdf_ready = bool(_sanitize_excel_text(record.source_pdf_path)) and Path(record.source_pdf_path).is_file()
    image_ready = bool(_sanitize_excel_text(record.local_image_path)) and Path(record.local_image_path).is_file()
    password_ready = bool(_sanitize_excel_text(record.password)) or source_pdf_ready
    return _compute_channel_status(
        contact_value=record.parent_email,
        sent_at=record.email_sent_at,
        error=record.email_error,
        note=record.note,
        asset_ready=password_ready and image_ready and bool(_sanitize_excel_text(record.order_url)),
    )


def write_child_records(
    job_folder: str | Path,
    records: Sequence[ChildRecord],
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Path:
    if _db_mode(db, workflow_item_id):
        _write_child_records_to_db(db, int(workflow_item_id or 0), records, disk_name=disk_name)
        return child_info_path(job_folder)

    path = child_info_path(job_folder)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = VISIBLE_SHEET_TITLE
    hidden = workbook.create_sheet(INTERNAL_SHEET_TITLE)
    hidden.sheet_state = "hidden"

    sheet.append(VISIBLE_HEADERS)
    hidden.append(INTERNAL_HEADERS)

    for record in records:
        record.mms_status = _compute_mms_status(record)
        record.email_status = _compute_email_status(record)
        sheet.append(
            [
                _sanitize_excel_text(record.child_name),
                _sanitize_excel_text(record.class_name),
                _sanitize_excel_text(record.parent_phone),
                _sanitize_excel_text(record.parent_email),
                _sanitize_excel_text(record.mms_status),
                _sanitize_excel_text(record.email_status),
                _sanitize_excel_text(record.note),
            ]
        )
        hidden.append(
            [
                _sanitize_excel_text(record.page_index),
                _sanitize_excel_text(record.password),
                _sanitize_excel_text(record.preview_url),
                _sanitize_excel_text(record.local_image_path),
                _sanitize_excel_text(record.source_pdf_path),
                _sanitize_excel_text(record.order_url),
                _sanitize_excel_text(record.object_key),
                _sanitize_excel_text(record.mms_sid),
                _sanitize_excel_text(record.mms_sent_at),
                _sanitize_excel_text(record.mms_error),
                _sanitize_excel_text(record.email_message_id),
                _sanitize_excel_text(record.email_sent_at),
                _sanitize_excel_text(record.email_error),
            ]
        )

    sheet.freeze_panes = "A2"
    widths = {
        1: 28,
        2: 18,
        3: 18,
        4: 28,
        5: 12,
        6: 12,
        7: 36,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _same_child_record(left: ChildRecord, right: ChildRecord) -> bool:
    if int(left.db_id or 0) > 0 and int(right.db_id or 0) > 0:
        return int(left.db_id) == int(right.db_id)
    left_pdf = _sanitize_excel_text(left.source_pdf_path)
    right_pdf = _sanitize_excel_text(right.source_pdf_path)
    if left_pdf or right_pdf:
        return left_pdf == right_pdf and int(left.page_index or 0) == int(right.page_index or 0)
    return (
        _record_key(left.child_name, left.class_name) == _record_key(right.child_name, right.class_name)
        and int(left.page_index or 0) == int(right.page_index or 0)
    )


def delete_child_record(
    job_folder: str | Path,
    record: ChildRecord,
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> bool:
    if _db_mode(db, workflow_item_id) and int(record.db_id or 0) > 0:
        _ensure_parent_delivery_schema(db)
        with db.connect() as conn, conn.cursor() as cur:
            cur.execute(
                "DELETE FROM parent_delivery_contacts WHERE id=%s AND workflow_run_id=%s",
                (int(record.db_id), int(workflow_item_id or 0)),
            )
        return True

    records = read_child_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    kept = [current for current in records if not _same_child_record(current, record)]
    if len(kept) == len(records):
        return False
    write_child_records(job_folder, kept, db=db, workflow_item_id=workflow_item_id, disk_name=disk_name)
    return True


def summarize_records(records: Sequence[ChildRecord]) -> Dict[str, int]:
    summary = {
        "total": len(records),
        "mms_ready": 0,
        "email_ready": 0,
        "missing_phone": 0,
        "missing_email": 0,
        "needs_review": 0,
        "failed": 0,
        "mms_sent": 0,
        "email_sent": 0,
    }
    for record in records:
        record.mms_status = _compute_mms_status(record)
        record.email_status = _compute_email_status(record)
        if record.mms_status == STATUS_READY:
            summary["mms_ready"] += 1
        if record.email_status == STATUS_READY:
            summary["email_ready"] += 1
        if not _sanitize_excel_text(record.parent_phone):
            summary["missing_phone"] += 1
        if not _sanitize_excel_text(record.parent_email):
            summary["missing_email"] += 1
        if record.mms_status == STATUS_SENT:
            summary["mms_sent"] += 1
        if record.email_status == STATUS_SENT:
            summary["email_sent"] += 1
        if record.mms_status == STATUS_FAILED or record.email_status == STATUS_FAILED:
            summary["failed"] += 1
        if record.mms_status == STATUS_NEEDS_REVIEW or record.email_status == STATUS_NEEDS_REVIEW:
            summary["needs_review"] += 1
    return summary


def _dedupe_contacts(entries: Iterable[ContactEntry]) -> list[ContactEntry]:
    seen: set[tuple[str, str, str]] = set()
    result: list[ContactEntry] = []
    for entry in entries:
        key = (_normalize_name_key(entry.name), entry.phone, entry.email)
        if key in seen:
            continue
        seen.add(key)
        result.append(entry)
    return result


def _pick_name_candidate(parts: Sequence[str]) -> str:
    candidates: list[str] = []
    for part in parts:
        cleaned = _clean_contact_name(part)
        if not cleaned:
            continue
        if cleaned.casefold() in GENERIC_CONTACT_FIELDS:
            continue
        if EMAIL_RE.search(cleaned) or PHONE_RE.search(cleaned):
            continue
        if re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", cleaned):
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?", cleaned):
            continue
        if len(cleaned) < 2:
            continue
        candidates.append(cleaned)
    if not candidates:
        return ""
    candidates.sort(key=lambda value: (-len(value), value.casefold()))
    return candidates[0]


def _contact_from_parts(parts: Sequence[str], *, fallback_name: str = "") -> Optional[ContactEntry]:
    values = [_sanitize_excel_text(part) for part in parts if _sanitize_excel_text(part)]
    if not values:
        return None
    combined = " | ".join(values)
    email = _extract_email(combined)
    phone = _extract_phone(combined)
    if not email and not phone:
        return None
    name = _pick_name_candidate(values) or _clean_contact_name(fallback_name)
    if not name:
        return None
    return ContactEntry(name=name, phone=phone, email=email)


def _normalize_contact_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", _sanitize_excel_text(value).casefold())


def _contact_header_words(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _sanitize_excel_text(value).casefold()).strip()


_CONTACT_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "full_name": (
        "name",
        "full name",
        "student name",
        "child name",
        "student full name",
        "child full name",
        "student",
        "child",
    ),
    "first_name": (
        "first",
        "first name",
        "firstname",
        "given name",
        "student first",
        "student first name",
        "child first",
        "child first name",
    ),
    "last_name": (
        "last",
        "last name",
        "lastname",
        "surname",
        "family name",
        "student last",
        "student last name",
        "child last",
        "child last name",
    ),
    "phone": (
        "phone",
        "phone number",
        "mobile",
        "cell",
        "cell phone",
        "telephone",
        "tel",
        "parent phone",
        "student phone",
        "student phone mobile",
        "contact phone",
        "guardian phone",
        "mother phone",
        "father phone",
    ),
    "email": (
        "email",
        "e mail",
        "email address",
        "parent email",
        "contact email",
        "contact1 email",
        "contact 1 email",
        "guardian email",
        "mother email",
        "father email",
    ),
}


def _contact_header_role_scores(header: object) -> dict[str, int]:
    text = _contact_header_words(header)
    normalized = _normalize_contact_header(header)
    if not text:
        return {}

    scores: dict[str, int] = {}
    for role, aliases in _CONTACT_HEADER_ALIASES.items():
        best = 0
        for alias in aliases:
            alias_text = _contact_header_words(alias)
            alias_norm = _normalize_contact_header(alias)
            if normalized == alias_norm:
                best = max(best, 100)
            elif alias_norm and alias_norm in normalized:
                best = max(best, 92)
            else:
                best = max(best, int(fuzz.token_set_ratio(text, alias_text)))
        scores[role] = best

    if re.search(r"\be[-\s]*mail\b", text) or "email" in normalized:
        scores["email"] = max(scores.get("email", 0), 96)
    if any(token in normalized for token in ("phone", "mobile", "cell", "telephone")) or re.search(r"\btel\b", text):
        scores["phone"] = max(scores.get("phone", 0), 96)
    if re.search(r"\b(first|given)\b", text):
        scores["first_name"] = max(scores.get("first_name", 0), 94)
    if re.search(r"\b(last|surname|family)\b", text):
        scores["last_name"] = max(scores.get("last_name", 0), 94)
    if re.search(r"\b(full\s+name|student\s+name|child\s+name)\b", text):
        scores["full_name"] = max(scores.get("full_name", 0), 96)

    # Avoid "Contact1Email" being treated as a generic name/contact column.
    if scores.get("email", 0) >= 90:
        scores["full_name"] = min(scores.get("full_name", 0), 50)
    if scores.get("phone", 0) >= 90:
        scores["full_name"] = min(scores.get("full_name", 0), 50)
    if any(token in normalized for token in ("birth", "dob", "dateofbirth", "birthday")):
        return {}
    return {role: score for role, score in scores.items() if score >= 80}


def _best_contact_header_role(header: object) -> tuple[str, int]:
    scores = _contact_header_role_scores(header)
    if not scores:
        return "", 0
    role, score = max(scores.items(), key=lambda item: item[1])
    return role, score


def _build_contact_column_map(headers: Sequence[object]) -> tuple[dict[str, object], int]:
    role_candidates: dict[str, list[tuple[int, int]]] = {
        "full_name": [],
        "first_name": [],
        "last_name": [],
        "phone": [],
        "email": [],
    }
    for index, header in enumerate(headers):
        role, score = _best_contact_header_role(header)
        if role:
            role_candidates[role].append((score, index))

    for candidates in role_candidates.values():
        candidates.sort(reverse=True)

    mapping: dict[str, object] = {
        "full_name": role_candidates["full_name"][0][1] if role_candidates["full_name"] else None,
        "first_name": role_candidates["first_name"][0][1] if role_candidates["first_name"] else None,
        "last_name": role_candidates["last_name"][0][1] if role_candidates["last_name"] else None,
        "phone": [index for _, index in role_candidates["phone"]],
        "email": [index for _, index in role_candidates["email"]],
    }

    has_name = mapping["full_name"] is not None or (
        mapping["first_name"] is not None and mapping["last_name"] is not None
    )
    has_contact = bool(mapping["phone"]) or bool(mapping["email"])
    if not has_name or not has_contact:
        return mapping, 0

    score = 0
    for role in ("full_name", "first_name", "last_name", "phone", "email"):
        if role_candidates[role]:
            score += role_candidates[role][0][0]
    score += 25 * len(role_candidates["phone"][:2])
    score += 25 * len(role_candidates["email"][:2])
    return mapping, score


def _cell_value(row: Sequence[object], index: Optional[int]) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return _sanitize_excel_text(row[index])


def _first_extracted_value(row: Sequence[object], indexes: Sequence[int], extractor: Any) -> str:
    for index in indexes:
        value = extractor(_cell_value(row, index))
        if value:
            return value
    return ""


def _contact_from_mapped_row(mapping: dict[str, object], row: Sequence[object]) -> Optional[ContactEntry]:
    full_name_index = mapping.get("full_name")
    first_index = mapping.get("first_name")
    last_index = mapping.get("last_name")
    phone_indexes = [int(index) for index in mapping.get("phone", [])]
    email_indexes = [int(index) for index in mapping.get("email", [])]

    if isinstance(full_name_index, int):
        name = _clean_contact_name(_cell_value(row, full_name_index))
    else:
        name = _clean_contact_name(f"{_cell_value(row, first_index)} {_cell_value(row, last_index)}")

    phone = _first_extracted_value(row, phone_indexes, _extract_phone)
    email = _first_extracted_value(row, email_indexes, _extract_email)

    if not phone:
        phone = _extract_phone(" | ".join(_sanitize_excel_text(value) for value in row))
    if not email:
        email = _extract_email(" | ".join(_sanitize_excel_text(value) for value in row))
    if not name:
        fallback = _contact_from_parts([_sanitize_excel_text(value) for value in row])
        name = fallback.name if fallback is not None else ""

    if not name or (not phone and not email):
        return None
    return ContactEntry(name=name, phone=phone, email=email)


def _extract_structured_contacts_from_rows(rows: Sequence[Sequence[object]]) -> Optional[list[ContactEntry]]:
    best_header_index = -1
    best_mapping: dict[str, object] = {}
    best_score = 0
    for header_row_index, row in enumerate(rows[:30]):
        mapping, score = _build_contact_column_map(row)
        if score > best_score:
            best_score = score
            best_header_index = header_row_index
            best_mapping = mapping

    if best_header_index < 0 or best_score < 180:
        return None

    entries: list[ContactEntry] = []
    for data_row in rows[best_header_index + 1:]:
        entry = _contact_from_mapped_row(best_mapping, data_row)
        if entry is not None:
            entries.append(entry)
    return _dedupe_contacts(entries)


def _extract_contacts_from_xlsx(path: Path) -> list[ContactEntry]:
    workbook = load_workbook(path, data_only=True)
    entries: list[ContactEntry] = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        structured_entries = _extract_structured_contacts_from_rows(rows)
        if structured_entries:
            entries.extend(structured_entries)
            continue
        for row in rows:
            parts = [_sanitize_excel_text(value) for value in row if _sanitize_excel_text(value)]
            entry = _contact_from_parts(parts)
            if entry is not None:
                entries.append(entry)
    return _dedupe_contacts(entries)


def _extract_contacts_from_csv(path: Path) -> list[ContactEntry]:
    entries: list[ContactEntry] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = list(reader)
        structured_entries = _extract_structured_contacts_from_rows(rows)
        if structured_entries:
            return structured_entries
        for row in rows:
            parts = [_sanitize_excel_text(value) for value in row if _sanitize_excel_text(value)]
            entry = _contact_from_parts(parts)
            if entry is not None:
                entries.append(entry)
    return _dedupe_contacts(entries)


def _extract_contacts_from_text_lines(lines: Sequence[str]) -> list[ContactEntry]:
    entries: list[ContactEntry] = []
    previous_name = ""
    for raw_line in lines:
        line = _sanitize_excel_text(raw_line)
        if not line:
            continue
        entry = _contact_from_parts([line], fallback_name=previous_name)
        if entry is not None:
            entries.append(entry)
            previous_name = entry.name
            continue
        candidate_name = _pick_name_candidate([line])
        if candidate_name:
            previous_name = candidate_name
    return _dedupe_contacts(entries)


def _extract_contacts_from_txt(path: Path) -> list[ContactEntry]:
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    return _extract_contacts_from_text_lines(lines)


def _extract_contacts_from_pdf(path: Path) -> list[ContactEntry]:
    from pypdf import PdfReader

    lines: list[str] = []
    reader = PdfReader(str(path))
    for page in reader.pages:
        text = page.extract_text() or ""
        lines.extend(text.splitlines())
    return _extract_contacts_from_text_lines(lines)


def extract_contacts_from_file(path: str | Path) -> list[ContactEntry]:
    file_path = Path(path).resolve()
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_contacts_from_xlsx(file_path)
    if suffix == ".csv":
        return _extract_contacts_from_csv(file_path)
    if suffix == ".txt":
        return _extract_contacts_from_txt(file_path)
    if suffix == ".pdf":
        return _extract_contacts_from_pdf(file_path)
    raise RuntimeError(f"Unsupported contact file: {file_path}")


def _mark_duplicate_review(records: Sequence[ChildRecord], indexes: Sequence[int], imported_name: str) -> None:
    message = f'Needs review: multiple child rows match imported name "{_sanitize_excel_text(imported_name)}".'
    for index in indexes:
        records[index].note = _append_note(records[index].note, message, replace_prefixes=("Needs review:", "Review:"))


def _match_contact_to_record(contact: ContactEntry, records: Sequence[ChildRecord]) -> Optional[int]:
    name_key = _normalize_name_key(contact.name)
    if not name_key:
        return None

    exact_indexes = [idx for idx, record in enumerate(records) if _normalize_name_key(record.child_name) == name_key]
    if len(exact_indexes) == 1:
        return exact_indexes[0]
    if len(exact_indexes) > 1:
        _mark_duplicate_review(records, exact_indexes, contact.name)
        return None

    candidates = {_normalize_name_key(record.child_name): idx for idx, record in enumerate(records) if _normalize_name_key(record.child_name)}
    if not candidates:
        return None
    extracted = process.extract(
        name_key,
        list(candidates.keys()),
        scorer=fuzz.token_set_ratio,
        limit=2,
    )
    if not extracted:
        return None
    best_key, best_score, _ = extracted[0]
    second_score = extracted[1][1] if len(extracted) > 1 else 0
    if best_score >= 96 and (best_score - second_score) >= 4:
        return candidates.get(best_key)
    return None


def import_contact_file(
    job_folder: str | Path,
    contact_file: str | Path,
    *,
    overwrite: bool = False,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Dict[str, object]:
    records = read_child_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    if not records:
        raise RuntimeError("Parent Contacts file does not exist yet. Build it from Stage 3 PDFs first.")

    contacts = extract_contacts_from_file(contact_file)
    matched = 0
    skipped_existing = 0
    unmatched = 0
    phone_updated = 0
    email_updated = 0
    unmatched_names: list[str] = []

    for contact in contacts:
        index = _match_contact_to_record(contact, records)
        if index is None:
            unmatched += 1
            if contact.name:
                unmatched_names.append(contact.name)
            continue
        record = records[index]
        updated_any = False
        if contact.phone:
            phone_value = _sanitize_excel_text(contact.phone)
            if record.parent_phone and not overwrite:
                skipped_existing += 1
            else:
                if phone_value != record.parent_phone:
                    record.mms_sid = ""
                    record.mms_sent_at = ""
                    record.mms_error = ""
                record.parent_phone = phone_value
                phone_updated += 1
                updated_any = True
        if contact.email:
            email_value = _sanitize_excel_text(contact.email).lower()
            if record.parent_email and not overwrite:
                skipped_existing += 1
            else:
                if email_value != record.parent_email:
                    record.email_message_id = ""
                    record.email_sent_at = ""
                    record.email_error = ""
                record.parent_email = email_value
                email_updated += 1
                updated_any = True
        if updated_any:
            matched += 1
            record.note = _strip_note_prefix(record.note, ("Needs review:", "Review:"))

    write_child_records(job_folder, records, db=db, workflow_item_id=workflow_item_id, disk_name=disk_name)
    summary = summarize_records(records)
    return {
        "contacts_found": len(contacts),
        "matched": matched,
        "unmatched": unmatched,
        "needs_review": summary["needs_review"],
        "skipped_existing": skipped_existing,
        "phone_updated": phone_updated,
        "email_updated": email_updated,
        "unmatched_names": unmatched_names[:50],
    }


def import_phone_file(
    job_folder: str | Path,
    contact_file: str | Path,
    *,
    overwrite: bool = False,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Dict[str, object]:
    return import_contact_file(
        job_folder,
        contact_file,
        overwrite=overwrite,
        db=db,
        workflow_item_id=workflow_item_id,
        disk_name=disk_name,
    )


def _clear_channel_failure_note(note: str, channel_prefix: str) -> str:
    return _strip_note_prefix(note, (f"{channel_prefix} failed:",))


def send_ready_child_mms(
    job_folder: str | Path,
    twilio_settings: Dict[str, str],
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Dict[str, int]:
    records = read_child_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    sent = 0
    failed = 0
    skipped = 0

    for record in records:
        record.mms_status = _compute_mms_status(record)
        if record.mms_status != STATUS_READY:
            skipped += 1
            continue
        try:
            result = send_sms(
                account_sid=str(twilio_settings.get("account_sid") or ""),
                auth_token=str(twilio_settings.get("auth_token") or ""),
                from_number=str(twilio_settings.get("from_number") or ""),
                to_phone=record.parent_phone,
                body=DEFAULT_PARENT_MMS_BODY,
                media_urls=[record.preview_url] if record.preview_url else None,
            )
            record.mms_sid = _sanitize_excel_text(result.get("sid"))
            record.mms_sent_at = _now_stamp()
            record.mms_error = ""
            record.note = _clear_channel_failure_note(record.note, "MMS")
            sent += 1
        except Exception as exc:
            message = _sanitize_excel_text(exc)
            record.mms_error = message
            record.note = _append_note(record.note, f"MMS failed: {message}", replace_prefixes=("MMS failed:",))
            failed += 1
    write_child_records(job_folder, records, db=db, workflow_item_id=workflow_item_id, disk_name=disk_name)
    return {"sent": sent, "failed": failed, "skipped": skipped}


def _parent_email_subject(record: ChildRecord) -> str:
    child_name = _sanitize_excel_text(record.child_name)
    if child_name:
        return f"{child_name} school photo proofs are ready"
    return DEFAULT_PARENT_EMAIL_SUBJECT


def _password_for_record(record: ChildRecord) -> str:
    password = _sanitize_excel_text(record.password)
    if password:
        return password
    pdf_path = _sanitize_excel_text(record.source_pdf_path)
    if not pdf_path or not Path(pdf_path).is_file():
        return ""
    try:
        import fitz  # type: ignore

        with fitz.open(pdf_path) as doc:
            if doc.page_count:
                return _extract_page_password(doc.load_page(0).get_text("text"))
    except Exception:
        return ""
    return ""


def _parent_email_body(record: ChildRecord) -> str:
    child_name = _sanitize_excel_text(record.child_name) or "Your child"
    password = _password_for_record(record)
    lines = [
        "Hello,",
        "",
        f"{child_name}'s school photo proofs are ready.",
        "Use the View & Order Online button in this email to open the order page.",
    ]
    if password:
        lines.extend(["", f"Password: {password}"])
    lines.extend(["", "The image below is a quick preview of your child's proof page."])
    lines.extend(["", "Thank you."])
    return "\n".join(lines)


def _parent_email_html_body(record: ChildRecord, *, image_cid: str) -> str:
    child_name = html.escape(_sanitize_excel_text(record.child_name) or "Your child")
    order_url = html.escape(_sanitize_excel_text(record.order_url), quote=True)
    password = html.escape(_password_for_record(record))
    link_html = ""
    if order_url:
        link_html = (
            f'<p><a href="{order_url}" '
            'style="display:inline-block;background:#2563eb;color:#ffffff;text-decoration:none;'
            'padding:12px 18px;border-radius:8px;font-weight:700;">View &amp; Order Online</a></p>'
        )
    password_html = (
        '<p style="font-size:18px;margin:16px 0;"><strong>Password: '
        f'{password}</strong></p>'
        if password
        else ""
    )
    return f"""
    <div style="font-family:Arial,Helvetica,sans-serif;color:#111827;line-height:1.45;">
      <p>Hello,</p>
      <p><strong>{child_name}'s school photo proofs are ready.</strong></p>
      <p>Use the button below to open the order page.</p>
      {link_html}
      {password_html}
      <p>The image below is a quick preview of your child's proof page.</p>
      <p><img src="cid:{html.escape(image_cid, quote=True)}" alt="Proof order preview"
              style="max-width:640px;width:100%;height:auto;border:1px solid #d1d5db;border-radius:8px;"></p>
      <p>Thank you.</p>
    </div>
    """


def send_ready_parent_email(
    job_folder: str | Path,
    *,
    db: Any = None,
    workflow_item_id: Optional[int] = None,
    disk_name: str = "",
) -> Dict[str, int]:
    from folder_manager.proofing_online.order_import.gmail_client import (
        get_gmail_service,
        send_email_with_inline_assets,
    )

    records = read_child_records(job_folder, db=db, workflow_item_id=workflow_item_id)
    service = get_gmail_service()
    sent = 0
    failed = 0
    skipped = 0

    for record in records:
        record.email_status = _compute_email_status(record)
        if record.email_status != STATUS_READY:
            skipped += 1
            continue
        try:
            record.password = _password_for_record(record)
            if not record.password:
                raise RuntimeError("Password is missing for this child PDF page.")
            image_cid = f"proof_preview_{record.db_id or record.page_index or sent + failed + skipped}"
            result = send_email_with_inline_assets(
                service,
                to=record.parent_email,
                subject=_parent_email_subject(record),
                body_text=_parent_email_body(record),
                html_body=_parent_email_html_body(record, image_cid=image_cid),
                inline_image_path=record.local_image_path,
                inline_image_cid=image_cid,
                attachment_paths=[],
            )
            record.email_message_id = _sanitize_excel_text(result.get("id"))
            record.email_sent_at = _now_stamp()
            record.email_error = ""
            record.note = _clear_channel_failure_note(record.note, "Email")
            sent += 1
        except Exception as exc:
            message = _sanitize_excel_text(exc)
            record.email_error = message
            record.note = _append_note(record.note, f"Email failed: {message}", replace_prefixes=("Email failed:",))
            failed += 1
    write_child_records(job_folder, records, db=db, workflow_item_id=workflow_item_id, disk_name=disk_name)
    return {"sent": sent, "failed": failed, "skipped": skipped}


__all__ = [
    "CHILD_ASSETS_DIRNAME",
    "CHILD_INFO_FILENAME",
    "DEFAULT_MMS_BODY",
    "DEFAULT_PARENT_EMAIL_SUBJECT",
    "DEFAULT_PARENT_MMS_BODY",
    "PARENT_CONTACTS_DISPLAY_NAME",
    "ChildRecord",
    "child_assets_path",
    "child_info_path",
    "delete_child_record",
    "extract_contacts_from_file",
    "import_contact_file",
    "import_phone_file",
    "prepare_child_info_assets",
    "read_child_records",
    "send_ready_child_mms",
    "send_ready_parent_email",
    "summarize_records",
    "validate_parent_email",
    "validate_parent_phone",
    "write_child_records",
]
